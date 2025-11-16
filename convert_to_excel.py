import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the markdown file
md_file_path = r"c:\BEBA_HOTEL\My API Project\‏‏Master Application Hotel Fixed\CustomerApi\HOTEL_CODES_LIST.md"
excel_file_path = r"c:\BEBA_HOTEL\My API Project\‏‏Master Application Hotel Fixed\CustomerApi\HOTEL_CODES_LIST.xlsx"

# Read markdown file
with open(md_file_path, 'r', encoding='utf-8') as f:
    content = f.read()

# Parse the table
lines = content.split('\n')
table_data = []
in_table = False
header_found = False

for line in lines:
    line = line.strip()
    if line.startswith('| # |') and 'Hotel Code' in line:
        # Header row
        header_found = True
        continue
    elif header_found and line.startswith('|---'):
        # Separator row
        in_table = True
        continue
    elif in_table and line.startswith('|') and not line.startswith('|---'):
        # Data row
        # Remove leading and trailing |
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) >= 5:
            # Clean up the cells (remove markdown formatting)
            row_data = []
            for cell in cells:
                # Remove backticks and extra spaces
                cell = re.sub(r'`([^`]+)`', r'\1', cell)
                cell = cell.strip()
                row_data.append(cell)
            table_data.append(row_data)
    elif in_table and not line.startswith('|'):
        # End of table
        break

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hotel Codes List"

# Set headers
headers = ["#", "Hotel Code", "Hotel Name", "Status", "Notes"]
ws.append(headers)

# Style the header row
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

# Add data rows
for row_data in table_data:
    ws.append(row_data)

# Style data rows
data_font = Font(size=11)
for row_num in range(2, len(table_data) + 2):
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style
        # Alternate row colors
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Auto-adjust column widths
column_widths = {
    'A': 8,   # #
    'B': 15,  # Hotel Code
    'C': 25,  # Hotel Name
    'D': 12,  # Status
    'E': 20   # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Set row height for header
ws.row_dimensions[1].height = 25

# Freeze header row
ws.freeze_panes = 'A2'

# Save the Excel file
wb.save(excel_file_path)
print(f"Excel file created successfully: {excel_file_path}")
print(f"Total hotels: {len(table_data)}")





